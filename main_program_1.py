"""               AUTO CREDIT
        Excel to PDF credit list generator
      Written and maintained by Thomas Philip
                    -lockdown project 2020
"""
# the program directory contains the python file and two text files for saving paths (excel_loc.txt , save_loc.txt)
import os
import tkinter as tk

""" The main function to work with excel file, generate credit list and converting to pdf """
def prgrm():
    with open("excel_loc.txt","r") as efile: # excel_loc.txt is a file saved in the same directory as the python program, it stores the source pagth
        path = efile.read()
    with open("save_loc.txt","r") as sfile: # save_loc.txt is a file saved in the same directory as the python program, it stores the saving pagth
        savepath = sfile.read()

    source_excel_path=os.path.normpath(path) # conerts path to native windows type

    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    import win32com.client   #import win32api for pdf conversion
    import datetime

    import locale # for printing amount in indian notation
    locale.setlocale(locale.LC_ALL, 'en_IN.utf8') # set locale on the computer if not working

    # create a new Excel workbook to store the results
    result_workbook = openpyxl.Workbook()
    # create a new excel sheet inside result_workbook
    result_sheet = result_workbook.active   # the result sheet stores the result data

    # in the result sheet, merge cells from A1 to H1 (row 1, col 1-8)
    result_sheet.merge_cells('A1:H1') # (i.e. a rectangular field is created) for the main title
    # find the current date for title
    current_date = datetime.datetime.today().date()
    # main title for the result sheet >>>  "PENDING PAYMENTS : dd-mm-yyyy"
    result_sheet.cell(row = 1, column = 1).value = 'PENDING PAYMENTS : '+ current_date.strftime('%d-%m-%Y') #(title is in the 1st row)

    # give the column names for the result sheet   (column names are in the 2nd row)
    # Date, Inv No, Retailor Name, Address, Amount, Disc., Received, PENDING  :(total 8 columns)
    result_sheet.cell(row = 2, column = 1).value="Date"
    result_sheet.cell(row = 2, column = 2).value="Inv.No"
    result_sheet.cell(row = 2, column = 3).value="Retailor Name"
    result_sheet.cell(row = 2, column = 4).value="Address"
    result_sheet.cell(row = 2, column = 5).value="Amount"
    result_sheet.cell(row = 2, column = 6).value="Disc."
    result_sheet.cell(row = 2, column = 7).value="Received"
    result_sheet.cell(row = 2, column = 8).value="PENDING"

    """ Get the location of the source excel file """
    # load the source excel workbook from the given source path
    sorce_workbook = openpyxl.load_workbook(source_excel_path)
    # store all the sheet names of the source workbook to a list (source_sheetnames)
    source_sheetnames=sorce_workbook.sheetnames # this list contains sheetnames A to Z

    # rn changed to result_current_row
    result_current_row = 3  # inside the result sheet, the rows used for storing the cell values starts from 3,
    # since rows 1 & 2 are used for main title & column headings

    # create a variable to keep track of the total pending amount
    result_total_pending = 0 # initialize it to 0

    """ code below is specific for the excel file im using """
    # for amt and rmax code, it is to know from which row there is no str fields in received column of each sheet. eg: 23450+3200 is str field
    rlst=[290,320,21,25,85,80,91,2,2,297,351,2,290,66,2,230,2,65,360,121,2,2,2,2,2,2] # 26 values for each sheet A to Z
    i=0 # for indexing rlst
    """ ............................................. """

    # (1st) for loop to go through each sheet inside the source workbook
    for character in source_sheetnames:
        source_sheet = sorce_workbook[character]  # sheets A to Z in source workbook is assigned to source_sheet after each iteration

        source_max_row = source_sheet.max_row # gets the maximum no. of rows in the current source_sheet

        # (2nd) for loop to iterate through each row from '2' to 'source_max_row' in the current source_sheet
        for source_current_row in range(rlst[i], source_max_row+1):  #source_current_row stores the no. of the current row
            """ above the range starts from rlst[i] instead of 2, to ensure that there is no strings or invalid type values in received column """
            """ try for a better approach for the above problem, try cleaning the source excel file to make sure the received column contains a single integer value"""

            # store the cell(column) date of the current row
            source_cell_date = source_sheet.cell(row = source_current_row, column = 1).value # date
            # check if date is not None and if so, store all the column values of the current row
            if source_cell_date != None :
                source_cell_inv_no = source_sheet.cell(row = source_current_row, column = 2).value # inv.no
                source_cell_retailor_name = source_sheet.cell(row = source_current_row, column = 3).value # Retailor name
                source_cell_address = source_sheet.cell(row = source_current_row, column = 4).value # Address
                source_cell_amount = source_sheet.cell(row = source_current_row, column = 5).value # Amount
                source_cell_discount = source_sheet.cell(row = source_current_row, column = 6).value # Discount
                # check if received column is None or not
                if  source_sheet.cell(row = source_current_row, column = 7).value == None:
                    source_cell_received = 0 # if received column is None, assign received = 0
                else:
                    source_cell_received = source_sheet.cell(row = source_current_row, column = 7).value # if not None, assign received it's column value
                # calculate the pending amount (amount - received)
                source_cell_pending = source_cell_amount - source_cell_received

                # to check whether pending amount is greater than 0(or any other min. amount (>0) we want)
                # only if it is True, we move all the stored column values of the current row in the source sheet to the result sheet
                if source_cell_pending > 0 :
                    # write column values to result sheet
                    result_sheet.cell(row = result_current_row, column = 1).value = source_cell_date.strftime('%d-%m-%Y') # change the Date to string format
                    # the date in the string format(if we store date in non-string format, it would result in gibberish) is moved to date column in the result sheet
                    result_sheet.cell(row = result_current_row, column = 2).value = source_cell_inv_no # Inv. no
                    result_sheet.cell(row = result_current_row, column = 3).value = source_cell_retailor_name # Retailor name
                    result_sheet.cell(row = result_current_row, column = 4).value = source_cell_address # Address
                    result_sheet.cell(row = result_current_row, column = 5).value = source_cell_amount # Amount
                    result_sheet.cell(row = result_current_row, column = 6).value = source_cell_discount # Discount
                    result_sheet.cell(row = result_current_row, column = 7).value = source_cell_received # Received
                    result_sheet.cell(row = result_current_row, column = 8).value = source_cell_pending # PENDING

                    # update total pending amount
                    result_total_pending = source_cell_pending + result_total_pending

                    # after writing all the columns in result sheet, increment the result_current_row for moving to the next row in result sheet
                    result_current_row += 1

        print(character)
        i += 1       # for rlst

    # move the total pending amount to the bottom of the result sheet
    # at the bottom of result sheet, merge cells from A(result_current_row) to H(result_current_row), i.e: (row: result_current_row, col: 1-8)
    result_total_pending_coordinates = ("A"+str(result_current_row)+":H"+str(result_current_row)) # the coordinates must be in the same string format as eg:('A45:H45')
    result_sheet.merge_cells(result_total_pending_coordinates) # (i.e. a rectangular field is created) for the total pending amount
    # writes the total pending amount in indian notation
    result_sheet.cell(row = result_current_row, column = 1).value = 'TOTAL PENDING AMOUNT : Rs '+ (locale.format_string("%d", result_total_pending, grouping=True))

    # print statement just for feedback if the whole extraction and writing process was successfully completed
    print("   SUCCESS ! ")

    """ below code is for formatting and styling the result sheet """
    # set the height of the 1st row for main title
    result_sheet.row_dimensions[1].height = 30
    # set the height of the 2nd row for column headings
    result_sheet.row_dimensions[2].height = 50
    # set the height of the current row(last row) for total pending amount
    result_sheet.row_dimensions[result_current_row].height = 30

    # set the width of each column ('A' means 1st column, 'B' 2nd column and so on...)
    result_sheet.column_dimensions['A'].width = 11  # Date column
    result_sheet.column_dimensions['B'].width = 10  # Inv. no column
    result_sheet.column_dimensions['C'].width = 30  # Retailor name column
    result_sheet.column_dimensions['D'].width = 20  # Address column
    result_sheet.column_dimensions['E'].width = 13  # Amount column
    result_sheet.column_dimensions['F'].width = 7   # Discount column
    result_sheet.column_dimensions['G'].width = 14  # Received column
    result_sheet.column_dimensions['H'].width = 14  # PENDING column

    # set the main title font style to bold
    result_sheet.cell(row = 1, column = 1).font = Font(size = 22, bold = True)
    # set the "total amount pending" font style to bold
    result_sheet.cell(row = result_current_row, column = 1).font = Font(size = 22, bold = True)

    # set font size for each coloumn
    result_sheet.cell(row = 2, column = 1).font = Font(size = 18) # Date column
    result_sheet.cell(row = 2, column = 2).font = Font(size = 18) # Inv. no column
    result_sheet.cell(row = 2, column = 3).font = Font(size = 18) # Retailor name column
    result_sheet.cell(row = 2, column = 4).font = Font(size = 18) # Address column
    result_sheet.cell(row = 2, column = 5).font = Font(size = 18) # Amount column
    result_sheet.cell(row = 2, column = 6).font = Font(size = 18) # Discount column
    result_sheet.cell(row = 2, column = 7).font = Font(size = 18) # Received column
    result_sheet.cell(row = 2, column = 8).font = Font(size = 18, bold = True, color='FF0000') # PENDING column, color is changed to red and font is BOLD

    # set centre alignment for main title
    title_alignment = Alignment(horizontal='center',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
    result_sheet.cell(row = 1, column = 1).alignment = title_alignment
    # set right alignment for "total amount pending"
    result_total_pending_alignment = Alignment(horizontal='right',vertical='bottom',text_rotation=0,wrap_text=False,shrink_to_fit=True,indent=0)
    result_sheet.cell(row = result_current_row, column = 1).alignment = result_total_pending_alignment

    """ Save the result workbook to result excel path """
    result_excel_path= os.path.normpath(savepath+"/Readymade_Creditlist.xlsx")
    # adds "/Readymade_Creditlist.xlsx" to the end of the savepath directory and creates a windows native path
    result_workbook.save(result_excel_path)
    # saves the creditlist excel file to the path

    """ Converting excel result sheet to PDF format """
    # set path for the pdf file
    # change path to native windows style having '\'
    result_pdf_path= os.path.normpath(savepath+"/Readymade_Creditlist.pdf")

    def print_excel_worksheet_to_pdf(i_sz_excel_path, i_sz_ws_name, i_sz_pdf_path):
        excel = win32com.client.Dispatch("Excel.Application")

        excel.Visible = False   #Keep the excel sheet closed
        excel.DisplayAlerts = False  #"Do you want to over write it?" Will not Pop up

        try:
            wb_source = excel.Workbooks.Open(i_sz_excel_path)

            ws_source = wb_source.Worksheets(i_sz_ws_name)
            ws_source.PageSetup.Orientation = 2 # change orientation to landscape, to fit all the columns in an A4 paper
            ws_source.Select()

            wb_source.ActiveSheet.ExportAsFixedFormat(0, i_sz_pdf_path)
        except Exception as e:
            print(e)

        excel.Application.Quit()
    #converts the result excel file to pdf file
    print_excel_worksheet_to_pdf(result_excel_path, 1, result_pdf_path)

    # print statement for feedback purpose
    print("PDF generated")

    # below code updates the gui status text box after creating the creditlist
    text_pdf.config(state="normal")
    text_pdf.delete("1.0", "end")
    text_pdf.insert(tk.END,"STATUS","center")
    text_pdf.insert(tk.END,"\n    Success... \n The Credit list has been created.","just_center")
    text_pdf.config(state='disabled')

#########################################################################################
""" GUI """

import tkinter as tk
from tkinter import filedialog as fd

window = tk.Tk()   # main parent window
window.title("AutoCredit") # title for the Application

with open("excel_loc.txt","r") as efile:
    path = efile.read()

with open("save_loc.txt","r") as sfile:
    savepath = sfile.read()

def selectfile():  # function for selecting the source excel file in gui
    path= fd.askopenfilename()
    with open("excel_loc.txt","w") as efile:
        efile.write(path)
    # loads the current file path and save path from the text files
    with open("excel_loc.txt","r") as efile:
        path = efile.read()
    with open("save_loc.txt","r") as sfile:
        savepath = sfile.read()
    # below code updates the file path label
    entry_file.config(state='normal')
    entry_file.delete(0, 'end')
    entry_file.insert(0,path)
    entry_file.config(state='disabled')
    # below code is for updating the status text box according the status of file path and save path
    if path=="" and savepath=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the Excel file \nand the location for saving credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='light green',bg='dark green')
        btn_change.config(fg='light green',bg='dark green')
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    elif path=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the Excel file.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='light green',bg='dark green')
        btn_change.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    elif savepath=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the location for saving\n the credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='black',bg="SystemButtonFace")
        btn_change.config(fg='light green',bg='dark green')
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    else :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease make sure to close the Excel file\nbefore generating the credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_pdf.config(state='normal')
        btn_open.config(fg='black',bg="SystemButtonFace")
        btn_change.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(fg='light green',bg='dark green')
        btn_pdf.config(state='normal')

def changeloc():  # function for selecting the save directory in gui
    savepath= fd.askdirectory()
    with open("save_loc.txt","w") as sfile:
        sfile.write(savepath)
    # loads the current file path and save path from the text files
    with open("excel_loc.txt","r") as efile:
        path = efile.read()
    with open("save_loc.txt","r") as sfile:
        savepath = sfile.read()
    # below code updates the file path label
    entry_save.config(state='normal')
    entry_save.delete(0, 'end')
    entry_save.insert(0,savepath)
    entry_save.config(state='disabled')
    # below code is for updating the status text box according the status of file path and save path
    if path=="" and savepath=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the Excel file \nand the location for saving credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='light green',bg='dark green')
        btn_change.config(fg='light green',bg='dark green')
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    elif path=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the Excel file.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='light green',bg='dark green')
        btn_change.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    elif savepath=="" :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease select the location for saving\n the credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_open.config(fg='black',bg="SystemButtonFace")
        btn_change.config(fg='light green',bg='dark green')
        btn_pdf.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(state='disabled')
    else :
        text_pdf.config(state='normal')
        text_pdf.delete(1.0, 'end')
        text_pdf.insert(tk.END,"STATUS","center")
        text_pdf.insert(tk.END,"\nPlease make sure to close the Excel file\nbefore generating the credit list.","just_center")
        text_pdf.config(state='disabled')
        btn_pdf.config(state='normal')
        btn_open.config(fg='black',bg="SystemButtonFace")
        btn_change.config(fg='black',bg="SystemButtonFace")
        btn_pdf.config(fg='light green',bg='dark green')
        btn_pdf.config(state='normal')

""" sets the layout of the main window and its frames in grid management method"""
window.rowconfigure(0, minsize=80, weight=1)
window.rowconfigure(1, minsize=50, weight=1)
window.rowconfigure(2, minsize=50, weight=1)
window.rowconfigure(3, minsize=50, weight=1)
window.rowconfigure(4, minsize=70, weight=1)

window.columnconfigure(0, minsize=500, weight=1)

label_title = tk.Label(window, text="AUTO CREDIT",fg = "light green",
		 bg = "dark green",
		 font = "Helvetica 20 bold italic")
fr_file = tk.Frame(window)
fr_file.grid(row=1,column=0)
fr_save = tk.Frame(window)
fr_save.grid(row=2,column=0)
fr_pdf = tk.Frame(window,highlightbackground="light green" ,highlightthickness=10)
fr_pdf.grid(row=3,column=0,pady=15)
fr_footer = tk.Frame(window)
fr_footer.grid(row=4,column=0)

lalbel_file = tk.Label(fr_file, text = "Excel file :",font="Ariel 12")
entry_file = tk.Entry(fr_file, width=50,font="Ariel 11")
entry_file.insert(0,path)
entry_file.config(state='disabled')

btn_open = tk.Button(fr_file, text="Select File",font="Ariel 12", command=selectfile)

lalbel_save = tk.Label(fr_save, text = "  Save at :",font="Ariel 12")
entry_save = tk.Entry(fr_save, width=50,font="Ariel 11")
entry_save.insert(0,savepath)
entry_save.config(state='disabled')

btn_change = tk.Button(fr_save, text="Change Location",font="Ariel 12",command=changeloc)

btn_pdf = tk.Button(fr_pdf, text="Generate Credit List",font="Ariel 12 bold", height=1,width=16,
            fg='light green',bg='dark green', command=prgrm)

# below code is run initially for updating the status text box according to the file path and save path
if path=="" and savepath=="" :
    text_pdf=tk.Text(fr_pdf,width=40, height=5,font="Ariel 13",bg="light green")
    text_pdf.config(state='normal')
    text_pdf.tag_configure("center", justify='center',font = "Ariel 13 bold",foreground="green")
    text_pdf.insert(tk.END,"STATUS","center")
    text_pdf.tag_configure("just_center", justify='center')
    text_pdf.insert(tk.END,"\nPlease select the Excel file \nand the location for saving credit list.","just_center")
    text_pdf.config(state='disabled')
    btn_open.config(fg='light green',bg='dark green')
    btn_change.config(fg='light green',bg='dark green')
    btn_pdf.config(fg='black',bg="SystemButtonFace")
    btn_pdf.config(state='disabled')
elif path=="" :
    text_pdf=tk.Text(fr_pdf,width=40, height=5,font="Ariel 13",bg="light green")
    text_pdf.config(state='normal')
    text_pdf.tag_configure("center", justify='center',font = "Ariel 13 bold",foreground="green")
    text_pdf.insert(tk.END,"STATUS","center")
    text_pdf.tag_configure("just_center", justify='center')
    text_pdf.insert(tk.END,"\nPlease select the Excel file.","just_center")
    text_pdf.config(state='disabled')
    btn_open.config(fg='light green',bg='dark green')
    btn_change.config(fg='black',bg="SystemButtonFace")
    btn_pdf.config(fg='black',bg="SystemButtonFace")
    btn_pdf.config(state='disabled')
elif savepath=="" :
    text_pdf=tk.Text(fr_pdf,width=40, height=5,font="Ariel 13",bg="light green")
    text_pdf.config(state='normal')
    text_pdf.tag_configure("center", justify='center',font = "Ariel 13 bold",foreground="green")
    text_pdf.insert(tk.END,"STATUS","center")
    text_pdf.tag_configure("just_center", justify='center')
    text_pdf.insert(tk.END,"\nPlease select the location for saving\n the credit list.","just_center")
    text_pdf.config(state='disabled')
    btn_open.config(fg='black',bg="SystemButtonFace")
    btn_change.config(fg='light green',bg='dark green')
    btn_pdf.config(fg='black',bg="SystemButtonFace")
    btn_pdf.config(state='disabled')
else :
    text_pdf=tk.Text(fr_pdf,width=40, height=5,font="Ariel 13",bg="light green")
    text_pdf.config(state='normal')
    text_pdf.tag_configure("center", justify='center',font = "Ariel 13 bold",foreground="green")
    text_pdf.insert(tk.END,"STATUS","center")
    text_pdf.tag_configure("just_center", justify='center')
    text_pdf.insert(tk.END,"\nPlease make sure to close the Excel file\nbefore generating the credit list.","just_center")
    text_pdf.config(state='disabled')
    btn_open.config(fg='black',bg="SystemButtonFace")
    btn_change.config(fg='black',bg="SystemButtonFace")
    btn_pdf.config(fg='light green',bg='dark green')
    btn_pdf.config(state='normal')

label_author = tk.Label(fr_footer,text="Developed by Thomas Philip",font="Ariel 10")
label_company = tk.Label(fr_footer,text="M.C JACOB & SONS",anchor="e",width=40,font="Ariel 11 bold")
label_estd = tk.Label(fr_footer,text="ESTD. 1929    ",anchor="e",width=40,fg="dark green",font="Ariel 10 bold")

label_title.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

lalbel_file.grid(row=0, column=0, sticky="e", padx=5)
entry_file.grid(row=0, column=1, sticky="e", padx=5)
btn_open.grid(row=1,column=1,sticky='e',padx=5,pady=5)

lalbel_save.grid(row=0, column=0, sticky="e", padx=5)
entry_save.grid(row=0, column=1, sticky="e", padx=5)
btn_change.grid(row=1,column=1,sticky='e',padx=5,pady=5)

btn_pdf.grid(row=0, column=0,pady=10)
text_pdf.grid(row=1, column=0,pady=10,padx=20)

label_author.grid(row=0 ,column=0,padx=10)
label_company.grid(row=0 ,column=1,padx=10)
label_estd.grid(row=1 ,column=1,padx=10)

window.mainloop()
"""..............................  thats it ! .........................."""
